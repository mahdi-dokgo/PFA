from datetime import date
from io import BytesIO

import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def export_excel(filename, headers, rows, title=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title or 'Export'

    header_fill = PatternFill('solid', fgColor='1e2235')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_pdf(filename, title, headers, rows, subtitle=None):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph('<b>ApprovWise</b>', styles['Title']))
    elements.append(Paragraph(title, styles['Heading2']))
    if subtitle:
        elements.append(Paragraph(subtitle, styles['Normal']))
    elements.append(Spacer(1, 20))

    data = [headers] + rows
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e2235')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_bc_detail_pdf(bc):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=40, bottomMargin=40, leftMargin=50, rightMargin=50,
    )
    styles = getSampleStyleSheet()
    elements = []

    # ── En-tête ──────────────────────────────────────────────────────────────
    elements.append(Paragraph('<b>ApprovWise</b>', styles['Title']))
    elements.append(Paragraph(f'BON DE COMMANDE — {bc.code}', styles['Heading1']))
    elements.append(Paragraph(
        f"Date d'édition : {date.today().strftime('%d/%m/%Y')}   |   "
        f"Statut : {bc.get_status_display()}   |   "
        f"Livraison prévue : {bc.expected_date.strftime('%d/%m/%Y')}",
        styles['Normal'],
    ))
    elements.append(Spacer(1, 16))

    # ── Fournisseur ──────────────────────────────────────────────────────────
    elements.append(Paragraph('<b>Fournisseur</b>', styles['Heading3']))
    supplier = bc.supplier
    fournisseur_lines = [supplier.name]
    if supplier.contact_name:
        fournisseur_lines.append(f'Contact : {supplier.contact_name}')
    if supplier.email:
        fournisseur_lines.append(f'Email : {supplier.email}')
    if supplier.phone:
        fournisseur_lines.append(f'Tél : {supplier.phone}')
    for line in fournisseur_lines:
        elements.append(Paragraph(line, styles['Normal']))
    elements.append(Spacer(1, 16))

    # ── Lignes ───────────────────────────────────────────────────────────────
    line_headers = ['Article', 'Réf.', 'Unité', 'Qté', 'Prix unit. (DH)', 'Sous-total (DH)']
    line_rows = []
    for ln in bc.lines.select_related('article').all():
        subtotal = ln.quantity * ln.unit_price
        line_rows.append([
            ln.article.name,
            ln.article.reference,
            ln.article.unit,
            str(ln.quantity),
            f'{ln.unit_price:.2f}',
            f'{subtotal:.2f}',
        ])
    total_row = ['', '', '', '', 'TOTAL TTC', f'{bc.total_amount:.2f} DH']

    table_data = [line_headers] + line_rows + [total_row]
    col_widths = [170, 55, 45, 35, 90, 90]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e2235')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (3, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (2, -1), 'LEFT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#e5e7eb')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#1e2235')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 40))

    # ── Zone de signatures ───────────────────────────────────────────────────
    elements.append(Paragraph('<b>Signatures</b>', styles['Heading3']))
    elements.append(Spacer(1, 8))
    sig_data = [
        ['Responsable Achats', 'Fournisseur'],
        ['', ''],
        ['', ''],
        ['Signature & cachet', 'Signature & cachet'],
    ]
    sig_table = Table(sig_data, colWidths=[240, 240])
    sig_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BOX', (0, 0), (0, -1), 0.5, colors.HexColor('#9ca3af')),
        ('BOX', (1, 0), (1, -1), 0.5, colors.HexColor('#9ca3af')),
        ('LINEBELOW', (0, 0), (-1, 0), 0.5, colors.HexColor('#9ca3af')),
        ('ROWHEIGHT', (0, 1), (-1, 2), 50),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(sig_table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="BC_{bc.code}.pdf"'
    return response
